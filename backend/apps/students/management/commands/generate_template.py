from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import Workbook


class Command(BaseCommand):
    help = "Generate standardized Excel template for students, subjects, enrollments, and metrics."

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            default="srm_template.xlsx",
            help="Output filename (default: srm_template.xlsx)",
        )

    def handle(self, *args, **options):
        output_name = options["output"]
        output_path = Path(output_name).resolve()

        wb = Workbook()
        ws_students = wb.active
        ws_students.title = "students"
        ws_students.append(["student_id", "full_name", "email", "enrolled_at"])
        ws_students.append(["S20057001", "Asif Iqbal", "asif.iqbal@students.uni.edu", "2026-04-30"])

        ws_subjects = wb.create_sheet("subjects")
        ws_subjects.append(["code", "name", "trimester", "year"])
        ws_subjects.append(["MIS602", "Data Modelling and Database Design", 3, 2026])

        ws_enr = wb.create_sheet("enrollments")
        ws_enr.append(["student_id", "subject_code", "trimester", "year"])
        ws_enr.append(["S20057001", "MIS602", 3, 2026])

        ws_metrics = wb.create_sheet("metrics")
        ws_metrics.append(
            [
                "student_id",
                "subject_code",
                "trimester",
                "year",
                "week",
                "attendance_pct",
                "tutorial_submission_pct",
                "assessment_attempt_pct",
                "action_taken",
                "action_date",
                "action_notes",
                "still_potentially_at_risk",
                "final_status",
            ]
        )
        ws_metrics.append(["S20057001", "MIS602", 3, 2026, 4, 45, 50, 20, "Called student", "2026-05-01", "Guardian informed", "", "At Risk"])
        ws_metrics.append(["S20057001", "MIS602", 3, 2026, 8, 62, 75, 90, "", "", "", "false", "Recovered"])

        wb.save(output_path)
        self.stdout.write(self.style.SUCCESS(f"Template generated: {output_path}"))
