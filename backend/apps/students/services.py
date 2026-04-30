from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Iterable

from django.db import transaction

from openpyxl import load_workbook

from apps.metrics.models import WeeklyMetric

from .models import Enrollment, Student, Subject


@transaction.atomic
def create_student(*, student_id: str, full_name: str, email: str, enrolled_at: date) -> Student:
    return Student.objects.create(
        student_id=student_id, full_name=full_name, email=email, enrolled_at=enrolled_at
    )


@transaction.atomic
def upsert_student(*, student_id: str, full_name: str, email: str, enrolled_at: date) -> Student:
    obj, _ = Student.objects.update_or_create(
        student_id=student_id,
        defaults={"full_name": full_name, "email": email, "enrolled_at": enrolled_at},
    )
    return obj


@transaction.atomic
def create_subject(*, code: str, name: str, trimester: int, year: int) -> Subject:
    return Subject.objects.create(code=code, name=name, trimester=trimester, year=year)


@transaction.atomic
def create_enrollment(*, student: Student, subject: Subject, trimester: int, year: int) -> Enrollment:
    obj, _ = Enrollment.objects.get_or_create(
        student=student, subject=subject, trimester=trimester, year=year
    )
    return obj


def _row_iter(ws) -> Iterable[tuple]:
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []
    return rows


def _normalize_header(v: object) -> str:
    return str(v).strip().lower() if v is not None else ""


def _parse_date(value: object) -> date:
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        return date.fromisoformat(value.strip())
    raise ValueError("Invalid date value. Use YYYY-MM-DD.")


def _sheet_rows_by_header(ws) -> tuple[dict[str, int], list[tuple]]:
    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return {}, []
    mapping = {_normalize_header(h): idx for idx, h in enumerate(headers)}
    values = [row for row in rows if row and any(c is not None and str(c).strip() != "" for c in row)]
    return mapping, values


@transaction.atomic
def bulk_upload_students(file_bytes: bytes) -> dict:
    """
    Supports:
    1) single-sheet legacy format: student_id | full_name | email | enrolled_at
    2) multi-sheet template:
       - students(student_id, full_name, email, enrolled_at)
       - subjects(code, name, trimester, year)
       - enrollments(student_id, subject_code, trimester, year)
       - metrics(student_id, subject_code, trimester, year, week,
                 attendance_pct, tutorial_submission_pct, assessment_attempt_pct)
    """
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet_map = {ws.title.strip().lower(): ws for ws in wb.worksheets}

    if "students" not in sheet_map:
        # Backward-compatible path: use active sheet with old 4-column format.
        created = 0
        updated = 0
        for row in _row_iter(wb.active):
            if not row or row[0] is None:
                continue
            student_id, full_name, email, enrolled_at = row[:4]
            existed = Student.objects.filter(student_id=str(student_id)).exists()
            upsert_student(
                student_id=str(student_id),
                full_name=str(full_name),
                email=str(email),
                enrolled_at=_parse_date(enrolled_at),
            )
            if existed:
                updated += 1
            else:
                created += 1
        return {
            "created": created,
            "updated": updated,
            "subjectsCreated": 0,
            "subjectsUpdated": 0,
            "enrollmentsCreated": 0,
            "metricsCreated": 0,
            "metricsUpdated": 0,
        }

    students_created = 0
    students_updated = 0
    subjects_created = 0
    subjects_updated = 0
    enrollments_created = 0
    metrics_created = 0
    metrics_updated = 0

    students_ws = sheet_map["students"]
    students_h, students_rows = _sheet_rows_by_header(students_ws)
    required_students = {"student_id", "full_name", "email", "enrolled_at"}
    missing = required_students - set(students_h.keys())
    if missing:
        raise ValueError(f"students sheet missing columns: {', '.join(sorted(missing))}")

    for row in students_rows:
        sid = str(row[students_h["student_id"]]).strip()
        if not sid:
            continue
        full_name = str(row[students_h["full_name"]]).strip()
        email = str(row[students_h["email"]]).strip()
        enrolled_at = _parse_date(row[students_h["enrolled_at"]])
        existed = Student.objects.filter(student_id=sid).exists()
        upsert_student(
            student_id=sid,
            full_name=full_name,
            email=email,
            enrolled_at=enrolled_at,
        )
        if existed:
            students_updated += 1
        else:
            students_created += 1

    if "subjects" in sheet_map:
        subjects_h, subject_rows = _sheet_rows_by_header(sheet_map["subjects"])
        required_subjects = {"code", "name", "trimester", "year"}
        missing = required_subjects - set(subjects_h.keys())
        if missing:
            raise ValueError(f"subjects sheet missing columns: {', '.join(sorted(missing))}")
        for row in subject_rows:
            code = str(row[subjects_h["code"]]).strip()
            if not code:
                continue
            defaults = {
                "name": str(row[subjects_h["name"]]).strip(),
                "trimester": int(row[subjects_h["trimester"]]),
                "year": int(row[subjects_h["year"]]),
            }
            _, created = Subject.objects.update_or_create(code=code, defaults=defaults)
            if created:
                subjects_created += 1
            else:
                subjects_updated += 1

    if "enrollments" in sheet_map:
        enr_h, enr_rows = _sheet_rows_by_header(sheet_map["enrollments"])
        required_enr = {"student_id", "subject_code", "trimester", "year"}
        missing = required_enr - set(enr_h.keys())
        if missing:
            raise ValueError(f"enrollments sheet missing columns: {', '.join(sorted(missing))}")
        for row in enr_rows:
            sid = str(row[enr_h["student_id"]]).strip()
            code = str(row[enr_h["subject_code"]]).strip()
            if not sid or not code:
                continue
            student = Student.objects.filter(student_id=sid).first()
            subject = Subject.objects.filter(code=code).first()
            if not student or not subject:
                continue
            trimester = int(row[enr_h["trimester"]])
            year = int(row[enr_h["year"]])
            _, created = Enrollment.objects.get_or_create(
                student=student, subject=subject, trimester=trimester, year=year
            )
            if created:
                enrollments_created += 1

    if "metrics" in sheet_map:
        met_h, met_rows = _sheet_rows_by_header(sheet_map["metrics"])
        required_met = {
            "student_id",
            "subject_code",
            "trimester",
            "year",
            "week",
            "attendance_pct",
            "tutorial_submission_pct",
            "assessment_attempt_pct",
        }
        missing = required_met - set(met_h.keys())
        if missing:
            raise ValueError(f"metrics sheet missing columns: {', '.join(sorted(missing))}")
        for row in met_rows:
            sid = str(row[met_h["student_id"]]).strip()
            code = str(row[met_h["subject_code"]]).strip()
            if not sid or not code:
                continue
            student = Student.objects.filter(student_id=sid).first()
            subject = Subject.objects.filter(code=code).first()
            if not student or not subject:
                continue
            trimester = int(row[met_h["trimester"]])
            year = int(row[met_h["year"]])
            week = int(row[met_h["week"]])
            if week not in (4, 8):
                continue
            enrollment, _ = Enrollment.objects.get_or_create(
                student=student, subject=subject, trimester=trimester, year=year
            )
            defaults = {
                "attendance_pct": row[met_h["attendance_pct"]],
                "tutorial_submission_pct": row[met_h["tutorial_submission_pct"]],
                "assessment_attempt_pct": row[met_h["assessment_attempt_pct"]],
            }
            _, created = WeeklyMetric.objects.update_or_create(
                enrollment=enrollment, week=week, defaults=defaults
            )
            if created:
                metrics_created += 1
            else:
                metrics_updated += 1

    return {
        "created": students_created,
        "updated": students_updated,
        "subjectsCreated": subjects_created,
        "subjectsUpdated": subjects_updated,
        "enrollmentsCreated": enrollments_created,
        "metricsCreated": metrics_created,
        "metricsUpdated": metrics_updated,
    }
